"""Document format management: upload an Excel template, infer its
structure (which cells are labels vs. input targets), and confirm it via
the Governance Queue before it can be used to generate real documents.

This implements steps ①②⑥ of the flow Noritsugu described on 2026-07-05:
upload template → AI infers structure → human confirms via Governance →
save as a named, reusable "format" (e.g. "フォーマットA"). Steps ③④⑤⑦
(feeding real data into a confirmed format to generate a filled document)
are a separate, not-yet-built phase — see `generate_document` TODO below.

Design choice: rather than submitting one Governance proposal per detected
field (which could be dozens per template), the entire field mapping for
a template is submitted as a *single* proposal. A human reviews/approves
the whole structure at once, matching the "confirm once, reuse forever"
flow Noritsugu described — not a field-by-field back-and-forth.

A confirmed format's status is never stored redundantly — `get_format`/
`list_formats` always resolve status by reading the linked
`governance_store` record live, so there is exactly one place a format's
approval state can live.
"""
from __future__ import annotations

import io
import json
from datetime import datetime, timezone
from typing import Any, Optional
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from capability.domain import Capability, CapabilityStatus, ExecutionStatus, GovernanceLevel
from services.capability_instance import ensure_registered, registry as capability_registry
from services import governance_store, file_storage, record_store
from services.llm_client import generate_text

# 2026-07-06 (Web化準備、docs/architecture.md 14.23): ローカルファイルへの
# 保存から、record_store（定義情報）+ file_storage（実ファイル）を使った
# Supabase保存に切り替えた。クラウド上でサーバーが再起動しても、承認済み
# フォーマットの定義や生成済みファイルが消えないようにするため。
FORMATS_TABLE = "app_document_formats"
TEMPLATES_BUCKET = "document-templates"
GENERATED_DOCS_BUCKET = "generated-documents"

DOCUMENT_FORMAT_CAPABILITY = Capability(
    capability_id="document_format_structure_inference",
    name="Document Format Structure Inference",
    category="business",
    description=(
        "Infers the label/input-cell structure of an uploaded Excel "
        "template (e.g. a customer-specific delivery note format), so it "
        "can later be filled in automatically. A wrong guess here could "
        "misplace real business data in generated documents, so this "
        "always requires human confirmation regardless of confidence — "
        "it is intentionally not eligible for auto-approval."
    ),
    owner_team="AI OS",
    owner_user_id="system",
    team_id="ai-os",
    status=CapabilityStatus.DEPLOYED,
    version="1.0.0",
    supported_inputs=["template_file"],
    supported_outputs=["field_mappings"],
    required_context=[],
    governance_level=GovernanceLevel.MEDIUM,
)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _append_jsonl(record: dict[str, Any]) -> None:
    """名前はJSONL時代のまま残しているが、実体はrecord_store経由で
    Supabaseに保存する（呼び出し元を一切変更せずに済むように）。"""
    record_store.append_record(FORMATS_TABLE, record)


def _read_jsonl() -> list[dict[str, Any]]:
    """同上 — 名前はJSONL時代のまま、実体はSupabaseから読む。"""
    return record_store.read_all_records(FORMATS_TABLE)


def _latest_by_format_id() -> dict[str, dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for record in _read_jsonl():
        format_id = record.get("format_id")
        if format_id:
            latest[format_id] = record
    return latest


def infer_structure(worksheet) -> list[dict[str, Any]]:
    """Heuristic: a non-empty text cell is a "label" if the cell to its
    right (preferred) or below it is empty — that empty cell becomes the
    presumed input location for that label. Labels ending in "："/":"
    score higher confidence, since that's a strong signal of "this is a
    form field," not incidental text.

    Formula cells are explicitly excluded from being treated as labels
    (fixed 2026-07-05, found via real testing on a complex real-world
    invoice template): without `data_only=True`, `cell.value` for a
    formula cell is the formula *string itself* (e.g. "=SUM(L32:L33)"),
    which is a `str` and was passing the "is this a label?" check,
    polluting detected fields with formula text. `cell.data_type == "f"`
    is openpyxl's own signal for "this cell holds a formula" — checked
    first since it's authoritative; the `startswith("=")` check is a
    defensive fallback in case `data_type` isn't set as expected.

    Two-pass design (2026-07-06, fixing a real bug found while adding
    table-region support): a naive single pass that always prefers
    "right" over "below" per-cell misclassifies the *last* column of a
    detail table as `direction="right"` whenever nothing happens to sit
    to its right (a common case — a table's last column, e.g. 金額, often
    has empty space beyond it). That breaks table-header grouping (which
    groups by shared "below" direction) by excluding exactly the column
    that should have joined. Fixed by first identifying which *rows*
    qualify as table headers (>=2 candidate labels with an empty cell
    below them), then, for cells in those rows, preferring "below" over
    "right" — so every column of a detected table consistently points
    below, regardless of what happens to sit to its immediate right.
    """
    from collections import defaultdict

    candidates: list[dict[str, Any]] = []
    for row in worksheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if getattr(cell, "data_type", None) == "f" or cell.value.strip().startswith("="):
                continue
            label = cell.value.strip()
            if not label:
                continue
            right_cell = worksheet.cell(row=cell.row, column=cell.column + 1)
            below_cell = worksheet.cell(row=cell.row + 1, column=cell.column)
            # MergedCell（結合セルのうち左上以外の部分）には書き込めない
            # (openpyxlは AttributeError を投げる)。ここで除外しないと、タイトル
            # のような結合セルが誤って「入力先」やテーブルヘッダーとして検出され、
            # 生成時にクラッシュする（2026-07-06、実ファイルで発見）。
            right_writable = right_cell.value in (None, "") and not isinstance(right_cell, MergedCell)
            below_writable = below_cell.value in (None, "") and not isinstance(below_cell, MergedCell)
            candidates.append({
                "cell": cell,
                "label": label,
                "right_cell": right_cell,
                "below_cell": below_cell,
                "right_empty": right_writable,
                "below_empty": below_writable,
            })

    by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for c in candidates:
        by_row[c["cell"].row].append(c)
    table_header_rows = {
        row for row, cells in by_row.items()
        if sum(1 for c in cells if c["below_empty"]) >= 2
    }

    mappings: list[dict[str, Any]] = []
    for c in candidates:
        row_num = c["cell"].row
        prefer_below = row_num in table_header_rows and c["below_empty"]

        target = None
        direction = None
        if prefer_below:
            target, direction = c["below_cell"], "below"
        elif c["right_empty"]:
            target, direction = c["right_cell"], "right"
        elif c["below_empty"]:
            target, direction = c["below_cell"], "below"

        if target is None:
            continue

        label = c["label"]
        is_labelish = label.rstrip().endswith(("：", ":"))
        mappings.append({
            "field_name": label.rstrip("：:").strip(),
            "label_cell": c["cell"].coordinate,
            "input_cell": target.coordinate,
            "direction": direction,
            "confidence": 0.7 if is_labelish else 0.5,
        })
    return mappings


def detect_table_regions(mappings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Detect repeating detail-table sections (2026-07-06): a single-value
    form field is typically alone on its row, but a table header (e.g.
    品番 | カラー | サイズ | 数量, as seen on a real customer invoice with
    a variable-length line-item table) has *multiple* "below"-direction
    labels sharing the same row. Two or more such labels on the same row
    are treated as one table's header row.

    Mutates `mappings` in place, tagging each grouped label with a
    `table_id` so a single flat `field_mappings` list can still represent
    both single fields and table columns without a second, disconnected
    data structure — `field_mappings` stays the one place a human edits
    field names/cells (13.6/14.7's existing review-and-edit UI keeps
    working unchanged), while `table_regions` is purely derived metadata
    for the repeating-row generation logic.

    Returns the list of detected table regions:
    [{"table_id", "header_row", "columns": [{"field_name", "label_cell",
    "column_letter"}, ...]}, ...]
    """
    from collections import defaultdict
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

    by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for m in mappings:
        if m.get("direction") != "below":
            continue
        _, row_num = coordinate_from_string(m["label_cell"])
        by_row[row_num].append(m)

    table_regions: list[dict[str, Any]] = []
    for row_num, row_mappings in by_row.items():
        if len(row_mappings) < 2:
            continue
        columns = sorted(
            row_mappings,
            key=lambda m: column_index_from_string(coordinate_from_string(m["label_cell"])[0]),
        )
        table_id = f"table-row{row_num}"
        table_regions.append({
            "table_id": table_id,
            "header_row": row_num,
            "columns": [
                {
                    "field_name": c["field_name"],
                    "label_cell": c["label_cell"],
                    "column_letter": coordinate_from_string(c["label_cell"])[0],
                }
                for c in columns
            ],
        })
        for c in row_mappings:
            c["table_id"] = table_id

    return table_regions


def create_format(name: str, file_bytes: bytes) -> dict[str, Any]:
    """Upload a template, infer its structure, and submit it to the
    Governance Queue for confirmation. Returns the format record
    (status will be "QUEUED_FOR_REVIEW" until a human decides).
    """
    ensure_registered(DOCUMENT_FORMAT_CAPABILITY)

    format_id = f"fmt-{uuid4().hex[:8]}"
    file_storage.upload_file(TEMPLATES_BUCKET, f"{format_id}.xlsx", file_bytes)

    trace_id = f"docformat-{uuid4().hex[:8]}"
    execution = capability_registry.execute_capability(
        capability_id=DOCUMENT_FORMAT_CAPABILITY.capability_id,
        inputs={"name": name},
        user_id="system",
        project_id="",
        trace_id=trace_id,
    )

    try:
        workbook = load_workbook(io.BytesIO(file_bytes))
        worksheet = workbook.active
        field_mappings = infer_structure(worksheet)
        table_regions = detect_table_regions(field_mappings)
        avg_confidence = (
            sum(m["confidence"] for m in field_mappings) / len(field_mappings)
            if field_mappings else 0.0
        )
    except Exception as e:
        capability_registry.record_execution_result(
            execution_id=execution.execution_id, outputs={},
            status=ExecutionStatus.FAILED,
            error_message=str(e),
        )
        raise

    capability_registry.record_execution_result(
        execution_id=execution.execution_id,
        outputs={"fields_detected": len(field_mappings), "tables_detected": len(table_regions)},
        status=ExecutionStatus.COMPLETED,
    )

    approval = governance_store.submit_proposal(
        source_capability_id=DOCUMENT_FORMAT_CAPABILITY.capability_id,
        concept=f"帳票フォーマット構造確認: {name}",
        ai_hypothesis=json.dumps(field_mappings, ensure_ascii=False),
        confidence_score=avg_confidence,
        trace_id=trace_id,
        proposal_id=format_id,
        governance_level=DOCUMENT_FORMAT_CAPABILITY.governance_level.value,
    )

    record = {
        "format_id": format_id,
        "name": name,
        "template_storage_path": f"{format_id}.xlsx",
        "field_mappings": field_mappings,
        "table_regions": table_regions,
        "governance_approval_id": approval.approval_id,
        "trace_id": trace_id,
        "created_at": _now(),
    }
    _append_jsonl(record)
    return _resolve_status(record)


def _resolve_status(record: dict[str, Any]) -> dict[str, Any]:
    """Attach live status/decision info from governance_store rather than
    storing it redundantly in the format record itself."""
    approval = governance_store.get_approval(record.get("governance_approval_id", ""))
    resolved = dict(record)
    resolved["status"] = approval.get("status") if approval else "UNKNOWN"
    resolved["approver_id"] = approval.get("approver_id") if approval else None
    resolved["approval_reason"] = approval.get("approval_reason") if approval else None
    return resolved


def list_formats() -> list[dict[str, Any]]:
    return [_resolve_status(r) for r in _latest_by_format_id().values()]


def get_format(format_id: str) -> Optional[dict[str, Any]]:
    record = _latest_by_format_id().get(format_id)
    return _resolve_status(record) if record else None


def _record_field_mapping_corrections_as_learning(
    original_record: dict[str, Any], edited_mappings: list[dict[str, Any]]
) -> None:
    """Stage C of the Learning Domain integration (2026-07-06): when a
    human edits the AI's detected field mappings (renames a field, fixes
    an input cell, or removes a false positive) before approving a
    format, that correction is exactly the kind of signal Learning is
    meant to observe — "the AI guessed X, a human corrected it to Y."

    2026-07-15（14.103、Noritsuguの指定）: 以前は`classify_and_scope`+
    `apply_candidate`を呼び、GOVERNED分類としてGovernance承認キューに
    投入していた。しかし、承認してもされなくても、Policy Memory（承認
    記録）を読んで実際のAIロジック（document_format_structure_inference
    のヒューリスティック等）を変更する仕組みがどこにも実装されておらず、
    承認という儀式に実質的な意味が無いと判明した（Learningモジュール
    全体が「候補を記録する」以上の消費者を持たない設計になっている
    ため）。承認キューに入れる意味の無い儀式は撤去し、`create_candidate`
    だけを呼んで「AIの構造推測をこの内容で人間が修正した」という事実を
    記録する（Activity Feedにも記録される）に留める。分類・スコープ付与・
    承認は行わない — 実際にAIロジックを見直す消費者ができた時に、改めて
    設計し直す。

    Best-effort only: any failure here (e.g. the optional `learning`
    module being unavailable) is swallowed so a Learning-recording bug
    can never block the actual field-mapping edit a person is trying to
    save.
    """
    try:
        original_mappings = original_record.get("field_mappings", [])
        original_by_cell = {m["label_cell"]: m for m in original_mappings}
        edited_by_cell = {m["label_cell"]: m for m in edited_mappings}

        corrections: list[dict[str, Any]] = []
        for label_cell, original in original_by_cell.items():
            edited = edited_by_cell.get(label_cell)
            if edited is None:
                corrections.append({
                    "label_cell": label_cell,
                    "change": "removed",
                    "before": {"field_name": original["field_name"], "input_cell": original["input_cell"]},
                })
            elif (
                edited["field_name"] != original["field_name"]
                or edited["input_cell"] != original["input_cell"]
            ):
                corrections.append({
                    "label_cell": label_cell,
                    "change": "modified",
                    "before": {"field_name": original["field_name"], "input_cell": original["input_cell"]},
                    "after": {"field_name": edited["field_name"], "input_cell": edited["input_cell"]},
                })

        if not corrections:
            return

        from learning import service as learning_service
        from learning.models import LearningSourceType

        learning_service.create_candidate(
            title=f"帳票フォーマット構造推測の人間による修正: {original_record.get('name', '')}",
            description=(
                f"「{original_record.get('name', '')}」のAI構造推測結果を、"
                f"人間が{len(corrections)}件修正しました(項目の削除・名称変更・入力先セルの修正)。"
            ),
            source_type=LearningSourceType.REPEATED_CORRECTION,
            created_by="document_format_review",
            confidence=0.6,
            evidence=corrections,
            suggested_application=(
                "document_format_structure_inference のヒューリスティックを、"
                "この修正パターン(evidence参照)を踏まえて見直す候補"
            ),
        )
    except Exception:
        pass


def update_field_mappings(format_id: str, field_mappings: list[dict[str, Any]]) -> dict[str, Any]:
    """Let a human directly edit the AI-detected field mappings (rename a
    field, fix which cell it points to, or remove a false positive like a
    misdetected formula or master-data cell) before confirming a format.

    Design (2026-07-06): rather than only offering a binary "trust the AI's
    guess entirely / reject it entirely" choice, this lets the human and
    the AI arrive at the structure together — the AI's guess is a starting
    point, not the final word. This appends a new record with the SAME
    `format_id` (the existing JSONL "latest record per format_id wins"
    read pattern in `_latest_by_format_id()` handles this for free) and
    keeps the same `governance_approval_id`, since editing-then-approving
    is one human review action, not a new AI proposal requiring a fresh
    Governance submission.

    Table regions are recomputed from the edited mappings (rather than
    edited directly) so they stay consistent with whatever the human
    changed — e.g. deleting a table-column row correctly shrinks that
    table's column list.

    Raises ValueError if the format doesn't exist.
    """
    record = _latest_by_format_id().get(format_id)
    if not record:
        raise ValueError(f"Format {format_id} not found")

    _record_field_mapping_corrections_as_learning(record, field_mappings)

    # table_id タグを一度リセットしてから再計算する(古いtable_idが残らないように)
    for m in field_mappings:
        m.pop("table_id", None)
    table_regions = detect_table_regions(field_mappings)

    updated = dict(record)
    updated["field_mappings"] = field_mappings
    updated["table_regions"] = table_regions
    updated["field_mappings_edited_at"] = _now()
    _append_jsonl(updated)
    return _resolve_status(updated)


INSTRUCTION_PARSING_CAPABILITY = Capability(
    capability_id="document_instruction_parsing",
    name="Chat Instruction Parsing for Document Fields",
    category="business",
    description=(
        "Parses a free-text chat instruction (e.g. '顧客名はUS_LOGS Inc.、"
        "数量は50個') into values for a confirmed format's named fields, "
        "using an LLM. This only pre-fills the generation form — it does "
        "not write anything to a real document by itself, and the human "
        "still reviews the filled-in values (and can edit them) before "
        "generating. No separate Governance approval is required per call "
        "for the same reason `document_generation` doesn't need one: the "
        "risky part (the format's structure) was already gated at "
        "confirmation time."
    ),
    owner_team="AI OS",
    owner_user_id="system",
    team_id="ai-os",
    status=CapabilityStatus.DEPLOYED,
    version="1.0.0",
    supported_inputs=["format_id", "instruction"],
    supported_outputs=["field_values"],
    required_context=[],
    governance_level=GovernanceLevel.LOW,
)


def parse_instruction_to_fields(format_id: str, instruction: str) -> dict[str, Any]:
    """Ask an LLM to map a free-text instruction onto a confirmed format's
    known field names (single fields only — table row values are not yet
    parsed from chat instructions, see docs/architecture.md). Returns
    {"field_values": {field_name: value, ...}} — only fields the LLM
    found a clear value for are included; fields it isn't confident about
    are simply omitted rather than guessed, so a human filling in the
    form can see at a glance what still needs manual input.
    """
    fmt = get_format(format_id)
    if not fmt:
        raise ValueError(f"Format {format_id} not found")
    if fmt["status"] != "APPROVED":
        raise ValueError(
            f"Format {format_id} is not APPROVED (status={fmt['status']}) — "
            "confirm it via Governance before using chat instructions with it."
        )

    field_names = [
        m["field_name"] for m in fmt["field_mappings"] if not m.get("table_id")
    ]

    ensure_registered(INSTRUCTION_PARSING_CAPABILITY)
    trace_id = f"docinstr-{uuid4().hex[:8]}"
    execution = capability_registry.execute_capability(
        capability_id=INSTRUCTION_PARSING_CAPABILITY.capability_id,
        inputs={"format_id": format_id, "instruction": instruction},
        user_id="system",
        project_id="",
        trace_id=trace_id,
    )

    prompt = f"""以下の項目名のリストと、ユーザーの指示文があります。
指示文から読み取れる値を、対応する項目名に割り当ててください。

項目名のリスト:
{json.dumps(field_names, ensure_ascii=False)}

ユーザーの指示文:
{instruction}

出力は、項目名をキー、値を文字列とするJSONオブジェクトのみを返してください。
前置きや説明文は一切含めないでください。指示文から明確に読み取れない項目は含めないでください。
出力例: {{"顧客名": "US_LOGS Inc.", "数量": "50"}}"""

    try:
        raw_response = generate_text(prompt, max_tokens=1000)
    except Exception as e:
        capability_registry.record_execution_result(
            execution_id=execution.execution_id, outputs={},
            status=ExecutionStatus.FAILED, error_message=str(e),
        )
        raise

    cleaned = raw_response.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.strip("`")
        if cleaned.startswith("json"):
            cleaned = cleaned[4:]
        cleaned = cleaned.strip()

    try:
        field_values = json.loads(cleaned)
        if not isinstance(field_values, dict):
            field_values = {}
    except json.JSONDecodeError:
        field_values = {}

    # 検出済みの項目名以外は無視する(LLMが勝手に項目を作らないようにする)
    field_values = {k: v for k, v in field_values.items() if k in field_names}

    capability_registry.record_execution_result(
        execution_id=execution.execution_id,
        outputs={"fields_filled": len(field_values)},
        status=ExecutionStatus.COMPLETED,
    )

    return {"field_values": field_values, "trace_id": trace_id}



DOCUMENT_GENERATION_CAPABILITY = Capability(
    capability_id="document_generation",
    name="Document Generation from Confirmed Format",
    category="business",
    description=(
        "Fills a human-confirmed document format's template with real "
        "project data plus user-supplied data. Only usable on APPROVED "
        "formats — the risky part (guessing the template's structure) was "
        "already gated by Governance when the format was confirmed, so "
        "each individual generation does not require separate approval; "
        "it's a mechanical, auditable repetition of an already-reviewed "
        "structure."
    ),
    owner_team="AI OS",
    owner_user_id="system",
    team_id="ai-os",
    status=CapabilityStatus.DEPLOYED,
    version="1.0.0",
    supported_inputs=["format_id", "project_id", "user_data"],
    supported_outputs=["generated_document"],
    required_context=["purchase_orders"],
    governance_level=GovernanceLevel.LOW,
)

# Best-effort mapping from common Japanese field-label text to
# `ProjectData` attributes, used to auto-fill from real project data when
# `project_id` is given. Deliberately small and explicit rather than
# guessing — an unmapped field name simply isn't auto-filled and shows up
# in `missing_fields` instead of silently getting a wrong value.
_INTERNAL_FIELD_MAP: dict[str, str] = {
    "顧客名": "customer_name",
    "顧客": "customer_name",
    "仕入先": "supplier_name",
    "仕入先名": "supplier_name",
    "PO番号": "po_number",
    "案件名": "po_number",
    "金額": "po_amount",
    "売上金額": "sale_amount",
    "原価": "cost_amount",
    "出荷日": "actual_delivery_date",
    "納期": "po_required_delivery_date",
    "請求日": "invoice_date",
    "支払期日": "payment_due_date",
}


def _internal_data_for_project(project_id: str) -> dict[str, Any]:
    """Best-effort auto-fill from real Supabase data via ProjectService.
    Returns {} on any failure — internal auto-fill is a convenience, never
    a hard requirement (user-supplied data can always cover a field this
    can't)."""
    if not project_id:
        return {}
    try:
        from services.project_service import ProjectService
        aggregate = ProjectService().build_project_aggregate(project_id)
        if not aggregate:
            return {}
        result: dict[str, Any] = {}
        for field_name, attr in _INTERNAL_FIELD_MAP.items():
            value = getattr(aggregate.data, attr, None)
            if value is not None:
                result[field_name] = value
        return result
    except Exception:
        return {}


def generate_document(
    format_id: str,
    project_id: str = "",
    user_data: Optional[dict[str, Any]] = None,
    table_rows: Optional[dict[str, list[dict[str, Any]]]] = None,
) -> dict[str, Any]:
    """Fill an APPROVED format's template with real internal data (via
    `project_id`) merged with `user_data` (e.g. invoice/packing-list/
    shipping details the user provides directly), matched to the format's
    single-value field_mappings by field_name. User-supplied values take
    precedence over internal auto-fill on overlap.

    `table_rows` (2026-07-06) fills repeating detail-table sections
    (e.g. 品番/カラー/サイズ/数量 line items) detected in
    `fmt["table_regions"]`: `{table_id: [{field_name: value, ...}, ...]}`
    — each dict in the list becomes one row, written starting directly
    below the table's header row. Rows are written in order starting at
    `header_row + 1`; existing rows below the header (if any) are
    overwritten, not inserted — this fills a static-sized template
    (typical for a customer's fixed-layout delivery note), it does not
    resize the sheet.

    Raises ValueError if the format doesn't exist or isn't APPROVED —
    generating from an unconfirmed structure guess is refused outright,
    not just warned about.
    """
    fmt = get_format(format_id)
    if not fmt:
        raise ValueError(f"Format {format_id} not found")
    if fmt["status"] != "APPROVED":
        raise ValueError(
            f"Format {format_id} is not APPROVED (status={fmt['status']}) — "
            "confirm it via the Governance queue before generating documents from it."
        )

    data: dict[str, Any] = _internal_data_for_project(project_id)
    data.update(user_data or {})
    table_rows = table_rows or {}

    ensure_registered(DOCUMENT_GENERATION_CAPABILITY)
    trace_id = f"docgen-{uuid4().hex[:8]}"
    execution = capability_registry.execute_capability(
        capability_id=DOCUMENT_GENERATION_CAPABILITY.capability_id,
        inputs={
            "format_id": format_id, "project_id": project_id,
            "data_keys": list(data.keys()), "table_row_counts": {k: len(v) for k, v in table_rows.items()},
        },
        user_id="system",
        project_id=str(project_id),
        trace_id=trace_id,
    )

    try:
        template_bytes = file_storage.download_file(TEMPLATES_BUCKET, fmt["template_storage_path"])
        workbook = load_workbook(io.BytesIO(template_bytes))
        worksheet = workbook.active

        filled: list[str] = []
        missing: list[str] = []
        write_errors: list[str] = []
        for mapping in fmt["field_mappings"]:
            if mapping.get("table_id"):
                continue  # テーブル列は下のループで別処理する
            field_name = mapping["field_name"]
            if field_name in data:
                try:
                    worksheet[mapping["input_cell"]] = data[field_name]
                    filled.append(field_name)
                except AttributeError:
                    # 結合セル（左上以外）など、書き込み不可なセルを指している場合。
                    # 1項目の失敗で生成全体を止めない — 承認前の「入力先セル」欄で
                    # 人間が直せるよう、はっきり報告する（2026-07-06、実ファイルで発見）。
                    write_errors.append(f"{field_name}（セル {mapping['input_cell']} は書き込めません。結合セルの可能性があります）")
            else:
                missing.append(field_name)

        tables_written: dict[str, int] = {}
        for region in fmt.get("table_regions", []):
            rows = table_rows.get(region["table_id"], [])
            for row_index, row_data in enumerate(rows):
                target_row = region["header_row"] + 1 + row_index
                for column in region["columns"]:
                    value = row_data.get(column["field_name"])
                    if value in (None, ""):
                        continue
                    cell_ref = f"{column['column_letter']}{target_row}"
                    try:
                        worksheet[cell_ref] = value
                        filled.append(f"{column['field_name']}[{row_index}]")
                    except AttributeError:
                        write_errors.append(
                            f"{column['field_name']}[{row_index}]（セル {cell_ref} は書き込めません。結合セルの可能性があります）"
                        )
            tables_written[region["table_id"]] = len(rows)
            if not rows:
                for column in region["columns"]:
                    missing.append(f"{column['field_name']}（テーブル: {region['table_id']}）")

        output_id = f"doc-{uuid4().hex[:8]}"
        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        file_storage.upload_file(GENERATED_DOCS_BUCKET, f"{output_id}.xlsx", output_buffer.getvalue())
    except Exception as e:
        capability_registry.record_execution_result(
            execution_id=execution.execution_id, outputs={},
            status=ExecutionStatus.FAILED, error_message=str(e),
        )
        raise

    capability_registry.record_execution_result(
        execution_id=execution.execution_id,
        outputs={"filled_count": len(filled), "missing_count": len(missing)},
        status=ExecutionStatus.COMPLETED,
    )

    return {
        "output_id": output_id,
        "output_storage_path": f"{output_id}.xlsx",
        "format_id": format_id,
        "format_name": fmt["name"],
        "filled_fields": filled,
        "missing_fields": missing,
        "tables_written": tables_written,
        "write_errors": write_errors,
        "trace_id": trace_id,
    }