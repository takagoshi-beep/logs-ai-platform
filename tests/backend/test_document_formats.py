"""Tests for `backend/services/document_formats.py`.

Covers, in order, the real bugs found via manual testing this session
(docs/architecture.md 14.7/14.8) plus the full upload -> review/edit ->
approve -> generate lifecycle:
- formula cells misread as labels (14.7)
- the "last column of a table gets the wrong direction" bug (14.8)
- merged cells crashing generation (14.8)
- table-region multi-row generation (14.8)
- human edits to field_mappings feeding the Learning Domain (14.10)
"""
from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook

from services import document_formats as df


def _workbook_bytes(build) -> bytes:
    """Build a Workbook via `build(ws)`, return it as .xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    build(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# infer_structure: single-field detection
# ---------------------------------------------------------------------------

def test_infer_structure_detects_right_direction_field():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "顧客名："
    mappings = df.infer_structure(ws)
    assert len(mappings) == 1
    assert mappings[0]["field_name"] == "顧客名"
    assert mappings[0]["input_cell"] == "B1"
    assert mappings[0]["direction"] == "right"
    assert mappings[0]["confidence"] == pytest.approx(0.7)  # ends with "："


def test_infer_structure_detects_below_direction_when_right_is_occupied():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "顧客名"  # no trailing "：" -> lower confidence
    ws["B1"] = 12345  # non-string: "occupies" B1 without itself being a label candidate
    mappings = df.infer_structure(ws)
    assert len(mappings) == 1
    assert mappings[0]["input_cell"] == "A2"
    assert mappings[0]["direction"] == "below"
    assert mappings[0]["confidence"] == pytest.approx(0.5)


def test_infer_structure_skips_fully_occupied_cells():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "顧客名"
    ws["A2"] = 1  # below occupied
    ws["B1"] = 1  # right occupied
    assert df.infer_structure(ws) == []


def test_infer_structure_excludes_formula_cells():
    """Real bug found 2026-07-05 on a formula-heavy customer invoice
    template: without data_only=True, a formula cell's .value is the
    formula *string itself*, which used to pass the "is this a label?"
    check and pollute detected fields with formula text."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "税込金額"
    ws["B1"] = "=SUM(L32:L33)"
    mappings = df.infer_structure(ws)
    field_names = [m["field_name"] for m in mappings]
    assert "=SUM(L32:L33)" not in field_names
    # "税込金額" itself is still a valid label (right cell holds a formula,
    # so not "empty" -> falls through to checking below instead)
    assert any(m["field_name"] == "税込金額" for m in mappings)


# ---------------------------------------------------------------------------
# detect_table_regions
# ---------------------------------------------------------------------------

def test_table_region_groups_shared_header_row():
    wb = Workbook()
    ws = wb.active
    ws["A12"] = "品番"
    ws["B12"] = "カラー"
    ws["C12"] = "数量"
    mappings = df.infer_structure(ws)
    regions = df.detect_table_regions(mappings)
    assert len(regions) == 1
    assert regions[0]["header_row"] == 12
    assert [c["field_name"] for c in regions[0]["columns"]] == ["品番", "カラー", "数量"]
    assert all(m.get("table_id") == regions[0]["table_id"] for m in mappings)


def test_single_field_alone_on_its_row_is_not_a_table():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "顧客名："
    mappings = df.infer_structure(ws)
    assert df.detect_table_regions(mappings) == []
    assert "table_id" not in mappings[0]


def test_table_last_column_joins_even_with_empty_space_to_its_right():
    """Real bug found 2026-07-06: a naive single-pass heuristic always
    preferred 'right' over 'below' per-cell, so a table's last column
    (nothing to its right) got direction='right' and was excluded from
    its own table's grouping. Fixed with a two-pass approach: rows that
    qualify as table headers force 'below' for every candidate in that
    row, regardless of what's immediately to the right."""
    wb = Workbook()
    ws = wb.active
    ws["A12"] = "品番"
    ws["B12"] = "カラー"
    ws["C12"] = "数量"
    ws["D12"] = "金額"  # nothing to its right at all
    mappings = df.infer_structure(ws)
    regions = df.detect_table_regions(mappings)
    assert len(regions) == 1
    assert [c["field_name"] for c in regions[0]["columns"]] == ["品番", "カラー", "数量", "金額"]


def test_merged_title_cell_excluded_from_detection():
    """Real bug found 2026-07-06: a merged title cell (e.g. 納品書,
    merged across several columns) was getting chosen as an input target
    and, worse, misdetected as a second table header. openpyxl raises
    AttributeError writing to a MergedCell's non-anchor position."""
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("D1:G1")
    ws["D1"] = "納品書"
    mappings = df.infer_structure(ws)
    # 納品書 itself is a valid label; its right/below targets must not be
    # inside the merge (E1/F1/G1 are MergedCell, D2 is a normal empty cell)
    assert len(mappings) == 1
    assert mappings[0]["field_name"] == "納品書"
    assert mappings[0]["input_cell"] == "D2"
    assert mappings[0]["direction"] == "below"


# ---------------------------------------------------------------------------
# create_format: upload -> Governance submission
# ---------------------------------------------------------------------------

def test_create_format_submits_to_governance_as_queued():
    content = _workbook_bytes(lambda ws: ws.__setitem__("A1", "顧客名："))
    record = df.create_format(name="テストフォーマット", file_bytes=content)
    assert record["status"] == "QUEUED_FOR_REVIEW"
    assert record["name"] == "テストフォーマット"
    assert len(record["field_mappings"]) == 1
    assert record["governance_approval_id"]


def test_list_and_get_format_reflect_created_record():
    content = _workbook_bytes(lambda ws: ws.__setitem__("A1", "顧客名："))
    created = df.create_format(name="一覧テスト", file_bytes=content)

    formats = df.list_formats()
    assert any(f["format_id"] == created["format_id"] for f in formats)

    fetched = df.get_format(created["format_id"])
    assert fetched["name"] == "一覧テスト"


def test_get_format_returns_none_for_unknown_id():
    assert df.get_format("fmt-does-not-exist") is None


# ---------------------------------------------------------------------------
# update_field_mappings: human edits + Learning Domain integration
# ---------------------------------------------------------------------------

def test_update_field_mappings_persists_edits_and_recomputes_tables():
    def build(ws):
        ws["A1"] = "顧客名："
        ws["A12"] = "品番"
        ws["B12"] = "カラー"
    content = _workbook_bytes(build)
    created = df.create_format(name="編集テスト", file_bytes=content)

    edited = [m for m in created["field_mappings"]]
    # 「カラー」列を削除する(誤検出だったことにする)
    edited = [m for m in edited if m["field_name"] != "カラー"]
    # 「顧客名」の名称を変更する
    for m in edited:
        if m["field_name"] == "顧客名":
            m["field_name"] = "取引先名"

    updated = df.update_field_mappings(created["format_id"], edited)

    names = {m["field_name"] for m in updated["field_mappings"]}
    assert names == {"取引先名", "品番"}
    # カラーが消えたので、テーブル領域はもう1列だけの「テーブル」ではなくなる
    # (2列未満はテーブルと見なされない = detect_table_regionsの条件)
    assert updated["table_regions"] == []


def test_update_field_mappings_raises_for_unknown_format():
    with pytest.raises(ValueError):
        df.update_field_mappings("fmt-does-not-exist", [])


def test_editing_out_a_false_positive_creates_a_governed_learning_candidate():
    """The Learning Domain integration added 2026-07-06 (docs/
    architecture.md 14.14): a human removing/renaming an AI-detected
    field is exactly the "AI guessed X, human corrected to Y" signal
    Learning is meant to observe."""
    def build(ws):
        ws["A1"] = "顧客名："
        ws["A3"] = "ゴミ項目"
    content = _workbook_bytes(build)
    created = df.create_format(name="学習連携テスト", file_bytes=content)

    edited = [m for m in created["field_mappings"] if m["field_name"] != "ゴミ項目"]
    df.update_field_mappings(created["format_id"], edited)

    from learning.repository import get_candidate_repository
    candidates = get_candidate_repository().list(learning_type="governed")
    assert len(candidates) == 1
    assert "学習連携テスト" in candidates[0].title
    assert candidates[0].evidence[0]["change"] == "removed"
    assert candidates[0].evidence[0]["before"]["field_name"] == "ゴミ項目"


def test_editing_with_no_actual_changes_creates_no_learning_candidate():
    content = _workbook_bytes(lambda ws: ws.__setitem__("A1", "顧客名："))
    created = df.create_format(name="無変更テスト", file_bytes=content)

    # 全く同じ内容のまま送り返す(実質的な修正なし)
    df.update_field_mappings(created["format_id"], created["field_mappings"])

    from learning.repository import get_candidate_repository
    assert get_candidate_repository().list() == []


# ---------------------------------------------------------------------------
# generate_document
# ---------------------------------------------------------------------------

def _approve(format_id: str) -> None:
    from services import governance_store
    fmt = df.get_format(format_id)
    governance_store.decide(
        approval_id=fmt["governance_approval_id"],
        decision="APPROVED",
        approver_id="test",
        reason="test",
    )


def test_generate_document_refuses_unapproved_format():
    content = _workbook_bytes(lambda ws: ws.__setitem__("A1", "顧客名："))
    created = df.create_format(name="未承認テスト", file_bytes=content)
    with pytest.raises(ValueError, match="not APPROVED"):
        df.generate_document(created["format_id"], user_data={"顧客名": "テスト"})


def test_generate_document_fills_single_fields_from_user_data():
    content = _workbook_bytes(lambda ws: ws.__setitem__("A1", "顧客名："))
    created = df.create_format(name="単一項目生成テスト", file_bytes=content)
    _approve(created["format_id"])

    result = df.generate_document(created["format_id"], user_data={"顧客名": "US_LOGS Inc."})
    assert result["filled_fields"] == ["顧客名"]
    assert result["missing_fields"] == []

    wb = load_workbook(io.BytesIO(df.file_storage.download_file(df.GENERATED_DOCS_BUCKET, f"{result['output_id']}.xlsx")))
    assert wb.active["B1"].value == "US_LOGS Inc."


def test_generate_document_reports_missing_fields_without_failing():
    content = _workbook_bytes(lambda ws: ws.__setitem__("A1", "顧客名："))
    created = df.create_format(name="欠落項目テスト", file_bytes=content)
    _approve(created["format_id"])

    result = df.generate_document(created["format_id"], user_data={})
    assert result["missing_fields"] == ["顧客名"]
    assert result["filled_fields"] == []


def test_generate_document_writes_multiple_table_rows_below_header():
    def build(ws):
        ws["A12"] = "品番"
        ws["B12"] = "カラー"
        ws["C12"] = "数量"
    content = _workbook_bytes(build)
    created = df.create_format(name="複数行生成テスト", file_bytes=content)
    _approve(created["format_id"])

    table_id = created["table_regions"][0]["table_id"]
    result = df.generate_document(
        created["format_id"],
        user_data={},
        table_rows={
            table_id: [
                {"品番": "LUL157", "カラー": "IVORY", "数量": "92"},
                {"品番": "LUL157", "カラー": "BLACK", "数量": "313"},
            ]
        },
    )
    assert result["tables_written"][table_id] == 2

    wb = load_workbook(io.BytesIO(df.file_storage.download_file(df.GENERATED_DOCS_BUCKET, f"{result['output_id']}.xlsx")))
    ws = wb.active
    assert (ws["A13"].value, ws["B13"].value, ws["C13"].value) == ("LUL157", "IVORY", "92")
    assert (ws["A14"].value, ws["B14"].value, ws["C14"].value) == ("LUL157", "BLACK", "313")


def test_generate_document_reports_merged_cell_write_failure_without_crashing():
    """Defense-in-depth added alongside the detection-time fix: even if a
    human manually re-points a field at a merged cell during review, the
    whole request must not crash (2026-07-06)."""
    def build(ws):
        ws["A4"] = "顧客名："
        ws.merge_cells("D1:G1")
    content = _workbook_bytes(build)
    created = df.create_format(name="安全網テスト", file_bytes=content)

    edited = created["field_mappings"]
    edited[0]["input_cell"] = "E1"  # 結合範囲内、非アンカー
    df.update_field_mappings(created["format_id"], edited)
    _approve(created["format_id"])

    result = df.generate_document(created["format_id"], user_data={"顧客名": "テスト値"})
    assert result["filled_fields"] == []
    assert len(result["write_errors"]) == 1
    assert "E1" in result["write_errors"][0]
